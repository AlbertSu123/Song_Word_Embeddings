import pandas as pd
import numpy as np
import openpyxl
vectors  = ['vectors_2000.txt','vectors_2001.txt','vectors_2002.txt','vectors_2003.txt','vectors_2004.txt','vectors_2005.txt','vectors_2006.txt','vectors_2007.txt','vectors_2008.txt', 'vectors_2009.txt','vectors_2010.txt', 'vectors_2011.txt', 'vectors_2012.txt', 'vectors_2013.txt', 'vectors_2014.txt', 'vectors_2015.txt']
vectors = ['vectors_2009.txt','vectors_2010.txt', 'vectors_2011.txt', 'vectors_2012.txt', 'vectors_2013.txt', 'vectors_2014.txt', 'vectors_2015.txt']
_row = 1
workbook = openpyxl.Workbook()
sheet = workbook.active
for vector in vectors:
    print(vector)
    n = len(open(vector, encoding="utf-8").readlines())
    options = ['bottles', 'vodka', 'alcohol', 'beer', 'hennessey', 'marijuana', 'weed',
               'kush', 'cocaine', 'coke', 'heroin', 'crack', 'heroin', 'hallucinogens', 'LSD', 'PCP',
               'ecstasy', 'inhalants', 'methamphatamine', 'opiates', 'amphetamines', 'tranquilizers', 'benzo']

    drug_list = []

    important_words = ['fun', 'lit', 'party', 'dope', 'expensive',
                       'high', 'hype', 'illegal', 'overdose', 'dead', 'danger', 'addictive', 'use', 'powerful']
    k = len(important_words)
    not_found = important_words[:]
    important_word_vectors = {}

    # adds the important word vectors to the important_word_vectors_dictionary.
    # important_word_vectors[important_word_string] = 50 dimensional np vector that is the important_word vector
    with open(vector, encoding="utf-8") as file:
        for _ in range(n):
            line = file.readline().split()
            if line[0] in not_found:
                not_found.remove(line[0])
                vec = np.zeros(50)
                for i in range(1, len(line)):
                    vec[i-1] = float(line[i])
                important_word_vectors[line[0]] = vec

    drug_distances = {}

    # now drug_distances[drug_name_string] = distances_dict
    # distances_dict[important_word] gives the euclidean distance from the drug name
    with open(vector, encoding="utf-8") as file:
        for _ in range(n):
            line = file.readline().split()
            if line[0] in options:
                drug_vec = np.zeros(50)
                for i in range(1, len(line)):
                    drug_vec[i-1] = float(line[i])
                if line[0] in options:
                    distances = {}
                    for key, value in important_word_vectors.items():
                        distance = sum((drug_vec - value) ** 2)
                        distances[key] = distance
                    drug_distances[line[0]] = distances
    min_val = {} #closest words to a drug
    lowest_five = []
    drug = 'weed'
    for _ in important_words:
        lowest_five = [v for k, v in drug_distances[drug].items()]
    # lowest_five.sort()
    for i in lowest_five:
        for k, v in drug_distances[drug].items():
            if i == v:
                min_val[k] = v
    if _row == 1:
        column = 2
        for key, values in min_val.items():
            sheet.cell(row= _row, column = column, value=key)
            column += 1
    _row += 1
    column = 1
    print(vector)
    for key,value in min_val.items():
        column += 1
        sheet.cell(row=_row, column = column, value=value)
workbook.save(filename ="marijuana.xlsx")

    #TO DO: Run all these scripts for each popular drug and paste them into excel
    #switch drug for the name of the drug we want and change the file names to the correct year
